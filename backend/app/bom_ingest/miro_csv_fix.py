from __future__ import annotations

import argparse
import re
import sqlite3
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable
import xml.etree.ElementTree as ET

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ITEM_NUMBER_RE = re.compile(r"^(?P<module>[A-Z]{2,3})(?P<number>\d{3})(?P<suffix>[PA])$")
NUMBERED_CELL_RE = re.compile(
    r"^(?P<item_number>[A-Z]{2,3}\d{3}[PA])\s*:\s*(?P<name>.+)$"
)
MODULE_TYPED_CELL_RE = re.compile(
    r"^(?P<module>[A-Z]{2,3})\s+(?P<suffix>[PA])\s*:\s*(?P<name>.+)$"
)
TYPE_ONLY_CELL_RE = re.compile(r"^(?P<suffix>[PA])\s*:\s*(?P<name>.+)$")
QUANTITY_RE = re.compile(r"#\s*(\d+)\s*$")

DEFAULT_MODULE_CODES = {"PL", "DRY", "AEC", "DAC", "DS", "FM", "MS", "POW", "UN"}
BLOCKER_STATUSES = {"needs_review", "conflict"}
BLOCKER_REASONS = {"needs_review", "conflict", "missing_parent", "ambiguous_type", "ambiguous_module"}
TRIMMED_INVENTORY_COLUMNS = {"partnumber", "partname", "Future_10k_min", "Future_10k_max", "avg", "current_cost_proto"}

PROTECTED_TOKENS = {
    "h2": "H2",
    "o2": "O2",
    "aec": "AEC",
    "un": "UN",
    "dry": "DRY",
    "dac": "DAC",
    "ds": "DS",
    "fm": "FM",
    "ms": "MS",
    "pow": "POW",
    "ecu": "ECU",
    "can": "CAN",
    "pcb": "PCB",
    "ntc": "NTC",
    "ip": "IP",
    "epdm": "EPDM",
    "m4": "M4",
    "m5": "M5",
    "m6": "M6",
    "m8": "M8",
    "m10": "M10",
    "2pin": "2pin",
    "4pin": "4pin",
    "8pin": "8pin",
    "12pin": "12pin",
    "8mm": "8mm",
    "12mm": "12mm",
    "25mm": "25mm",
    "30mm": "30mm",
    "34mm": "34mm",
    "80amp": "80Amp",
    "c-tube": "C-Tube",
    "ctube": "C-Tube",
    "o-ring": "O-Ring",
    "oring": "O-Ring",
    "g1/4": "G1/4",
}


@dataclass(frozen=True)
class InventoryItem:
    partnumber: str
    partname: str
    module_code: str
    suffix: str
    normalized_name: str


@dataclass
class ReviewOverride:
    review_decision: str | None = None
    approved_module: str | None = None
    approved_suffix: str | None = None
    approved_item_name: str | None = None
    approved_quantity: int | None = None
    approved_comment: str | None = None
    approved_item_number: str | None = None


@dataclass
class ParsedCell:
    row_index: int
    tree_level: int
    raw_text: str
    cleaned_text: str
    item_text_without_comment: str
    comment: str
    quantity: int
    explicit_item_number: str | None
    explicit_module: str | None
    explicit_sequence_number: int | None
    explicit_suffix: str | None
    inferred_module: str | None
    inferred_suffix: str | None
    normalized_item_name: str
    parent_raw_text: str
    parent_resolved_item_number: str | None
    resolution_status: str
    resolved_item_number: str | None
    resolved_item_name: str | None
    action: str
    is_ambiguous: bool
    blocker_reason: str | None
    syntax_kind: str
    has_children: bool
    parent_key: str | None
    cell_key: str
    source_column: str
    review_override: ReviewOverride | None = None


class InventoryAuthority:
    def __init__(self, items: Iterable[InventoryItem], source_df: pd.DataFrame):
        self.items = list(items)
        self.source_df = source_df.copy()
        self.by_number = {item.partnumber: item for item in self.items}
        self.by_name: dict[str, list[InventoryItem]] = {}
        self.by_name_module_suffix: dict[tuple[str, str, str], list[InventoryItem]] = {}
        self.module_codes = set(DEFAULT_MODULE_CODES)

        for item in self.items:
            self.module_codes.add(item.module_code)
            self.by_name.setdefault(item.normalized_name, []).append(item)
            key = (item.normalized_name, item.module_code, item.suffix)
            self.by_name_module_suffix.setdefault(key, []).append(item)

    @classmethod
    def from_inventory_file(cls, path: Path) -> "InventoryAuthority":
        df = pd.read_excel(path)
        df = trim_inventory_dataframe(df)

        items: list[InventoryItem] = []
        for _, row in df.iterrows():
            raw_number = collapse_whitespace(row.get("partnumber", ""))
            raw_name = collapse_whitespace(row.get("partname", ""))
            match = ITEM_NUMBER_RE.match(raw_number)
            if not match or not raw_name:
                continue
            items.append(
                InventoryItem(
                    partnumber=raw_number,
                    partname=normalize_item_name(raw_name),
                    module_code=match.group("module"),
                    suffix=match.group("suffix"),
                    normalized_name=normalize_item_name(raw_name),
                )
            )
        return cls(items, df)

    def lookup_by_number(self, item_number: str) -> InventoryItem | None:
        return self.by_number.get(item_number)

    def lookup_by_name(self, normalized_name: str) -> list[InventoryItem]:
        return list(self.by_name.get(normalized_name, []))

    def lookup_exact(self, normalized_name: str, module_code: str, suffix: str) -> list[InventoryItem]:
        return list(self.by_name_module_suffix.get((normalized_name, module_code, suffix), []))

    def max_sequence_for_module(self, module_code: str) -> int:
        max_number = -1
        for item in self.items:
            if item.module_code != module_code:
                continue
            match = ITEM_NUMBER_RE.match(item.partnumber)
            if match:
                max_number = max(max_number, int(match.group("number")))
        return max_number


def collapse_whitespace(text: object) -> str:
    return re.sub(r"\s+", " ", str(text).strip())


def is_nonempty(value: object) -> bool:
    if pd.isna(value):
        return False
    return bool(collapse_whitespace(value))


def trim_inventory_dataframe(df_inventory: pd.DataFrame) -> pd.DataFrame:
    invalid_start = next((i for i, col in enumerate(df_inventory.columns) if str(col).strip() == ""), None)
    if invalid_start is not None:
        df_inventory = df_inventory.iloc[:, :invalid_start]

    unnamed_columns = [col for col in df_inventory.columns if str(col).startswith("Unnamed:")]
    if unnamed_columns:
        df_inventory = df_inventory.drop(columns=unnamed_columns)

    for required in ("partnumber", "partname"):
        if required not in df_inventory.columns:
            df_inventory[required] = pd.NA

    keep_columns = [col for col in df_inventory.columns if col in TRIMMED_INVENTORY_COLUMNS or col in {"partnumber", "partname"}]
    return df_inventory[keep_columns].copy()


def _smart_title_word(word: str) -> str:
    key = word.casefold()
    if key in PROTECTED_TOKENS:
        return PROTECTED_TOKENS[key]

    if "-" in word:
        return "-".join(_smart_title_word(part) for part in word.split("-"))

    if "/" in word and key in PROTECTED_TOKENS:
        return PROTECTED_TOKENS[key]

    if not word:
        return word
    return word[:1].upper() + word[1:].lower()


def normalize_item_name(name: str) -> str:
    text = collapse_whitespace(name)
    if not text:
        return ""

    text = re.sub(r"\bO[\s-]?Ring\b", "O-Ring", text, flags=re.IGNORECASE)
    text = re.sub(r"\bC[\s-]?Tube\b", "C-Tube", text, flags=re.IGNORECASE)

    normalized = " ".join(_smart_title_word(word) for word in text.split(" "))
    normalized = re.sub(r"\bAec\b", "AEC", normalized)
    normalized = re.sub(r"\bUn\b", "UN", normalized)
    normalized = re.sub(r"\bDry\b", "DRY", normalized)
    normalized = re.sub(r"\bCan\b", "CAN", normalized)
    normalized = re.sub(r"\bPcb\b", "PCB", normalized)
    normalized = re.sub(r"\bEcu\b", "ECU", normalized)
    normalized = re.sub(r"\bNtc\b", "NTC", normalized)
    normalized = re.sub(r"\bIp\b", "IP", normalized)
    normalized = re.sub(r"\bEpdm\b", "EPDM", normalized)
    normalized = re.sub(r"\bOring\b", "O-Ring", normalized, flags=re.IGNORECASE)
    return collapse_whitespace(normalized)


def is_material_name_difference(imported_name: str, authority_name: str) -> bool:
    return normalize_item_name(imported_name) != normalize_item_name(authority_name)


def default_inventory_path(repo_root: Path) -> Path:
    candidate_dirs = [repo_root / "database", repo_root / "ingestion data raw"]
    candidates: list[Path] = []
    for directory in candidate_dirs:
        if directory.exists():
            candidates.extend(sorted(directory.glob("*.xlsx")))
    if not candidates:
        raise FileNotFoundError("No inventory .xlsx file found in 'database' or 'ingestion data raw'.")
    if len(candidates) == 1:
        return candidates[0]

    preferred = [path for path in candidates if "inventory" in path.name.casefold()]
    if preferred:
        return preferred[0]
    return candidates[0]


def load_bom_csv(path: Path) -> pd.DataFrame:
    return pd.read_csv(path, dtype=object)


def load_bom_opml(path: Path) -> pd.DataFrame:
    tree = ET.parse(path)
    body = tree.getroot().find("body")
    if body is None:
        raise ValueError("OPML file is missing a <body> element.")

    outline_children = [child for child in body if child.tag == "outline"]
    if len(outline_children) != 1:
        raise ValueError("OPML input must have exactly one root <outline> under <body>.")

    rows: list[list[str]] = []

    def walk(node: ET.Element, path_parts: list[str]) -> None:
        text = collapse_whitespace(node.attrib.get("text", ""))
        if not text:
            return
        current_path = path_parts + [text]
        rows.append(current_path)
        for child in node:
            if child.tag == "outline":
                walk(child, current_path)

    walk(outline_children[0], [])
    max_depth = max(len(row) for row in rows)
    records = []
    columns = [str(i) for i in range(max_depth)]
    for row in rows:
        padded = row + [pd.NA] * (max_depth - len(row))
        records.append(dict(zip(columns, padded, strict=False)))
    return pd.DataFrame(records, columns=columns, dtype=object)


def load_bom_input(path: Path) -> tuple[pd.DataFrame, str]:
    suffix = path.suffix.lower()
    if suffix == ".opml":
        return load_bom_opml(path), "opml"
    return load_bom_csv(path), "csv"


def repair_mindmap_tree(df_mindmap: pd.DataFrame) -> pd.DataFrame:
    df_mindmap = df_mindmap.copy()
    mindmap_depth = len(df_mindmap.columns)
    df_assy_tree = df_mindmap.copy()

    for i in range(mindmap_depth):
        for j in range(len(df_assy_tree)):
            if i < mindmap_depth - 1:
                if pd.isna(df_assy_tree.iloc[j, i + 1]):
                    df_assy_tree.iloc[j, i] = np.nan
            else:
                df_assy_tree.iloc[j, i] = np.nan

            if pd.isna(df_assy_tree.iloc[j, i]):
                continue

    list_assy_add: list[int] = []
    df_add = pd.DataFrame()

    for i in reversed(range(mindmap_depth)):
        last_item_name = df_assy_tree.iloc[0, i]
        for j in range(len(df_assy_tree)):
            item_name = df_assy_tree.iloc[j, i]
            if item_name != last_item_name:
                if last_item_name != "" and not pd.isna(last_item_name):
                    add_row = df_assy_tree.iloc[j - 1].copy()
                    add_row.iloc[i + 1 :] = np.nan
                    list_assy_add.append(j)
                    df_add = pd.concat([df_add, pd.DataFrame([add_row])], ignore_index=True)
                last_item_name = item_name
        if last_item_name != "" and not pd.isna(last_item_name):
            add_row = df_assy_tree.iloc[j - 1].copy()
            add_row.iloc[i + 1 :] = np.nan
            list_assy_add.append(j + 1)
            df_add = pd.concat([df_add, pd.DataFrame([add_row])], ignore_index=True)

    if not df_add.empty:
        df_add["list_assy_add"] = list_assy_add
        df_add = df_add.sort_values(by="list_assy_add")
        list_assy_add = df_add["list_assy_add"].tolist()
        df_add = df_add.drop(columns=["list_assy_add"])

        for i in range(len(df_add)):
            new_row = df_add.iloc[i]
            top = df_mindmap.iloc[: list_assy_add[i] + i]
            bottom = df_mindmap.iloc[list_assy_add[i] + i :]
            df_mindmap = pd.concat([top, pd.DataFrame([new_row]), bottom], ignore_index=True)

    df_mindmap = df_mindmap.drop_duplicates()
    df_mindmap = df_mindmap.dropna(how="all")
    df_mindmap = df_mindmap.reset_index(drop=True)
    return df_mindmap


def _parse_explicit_structure(
    base_text: str, module_codes: set[str]
) -> tuple[str, str | None, str | None, int | None, str | None, str]:
    numbered_match = NUMBERED_CELL_RE.match(base_text)
    if numbered_match:
        item_number = numbered_match.group("item_number")
        number_parts = ITEM_NUMBER_RE.match(item_number)
        assert number_parts is not None
        return (
            "numbered",
            item_number,
            number_parts.group("module"),
            int(number_parts.group("number")),
            number_parts.group("suffix"),
            numbered_match.group("name"),
        )

    module_typed_match = MODULE_TYPED_CELL_RE.match(base_text)
    if module_typed_match and module_typed_match.group("module").upper() in module_codes:
        return (
            "module_typed",
            None,
            module_typed_match.group("module").upper(),
            None,
            module_typed_match.group("suffix").upper(),
            module_typed_match.group("name"),
        )

    type_only_match = TYPE_ONLY_CELL_RE.match(base_text)
    if type_only_match:
        return (
            "type_only",
            None,
            None,
            None,
            type_only_match.group("suffix").upper(),
            type_only_match.group("name"),
        )

    prefix_match = re.match(r"^(?P<module>[A-Za-z]{2,3})(?P<separator>\s*-\s*|\s+)(?P<name>.+)$", base_text)
    if prefix_match and prefix_match.group("module").upper() in module_codes:
        return (
            "module_prefix",
            None,
            prefix_match.group("module").upper(),
            None,
            None,
            prefix_match.group("name"),
        )

    return ("plain", None, None, None, None, base_text)


def _apply_review_override(cell: ParsedCell, override: ReviewOverride | None) -> None:
    if override is None:
        return

    cell.review_override = override
    cell.blocker_reason = None
    cell.is_ambiguous = False
    cell.resolution_status = "skipped"
    cell.action = "skip"
    if override.approved_item_name:
        cell.normalized_item_name = normalize_item_name(override.approved_item_name)
    if override.approved_quantity is not None:
        cell.quantity = int(override.approved_quantity)
    if override.approved_comment is not None:
        cell.comment = collapse_whitespace(override.approved_comment)
    if override.approved_module:
        cell.explicit_module = override.approved_module.upper()
    if override.approved_suffix:
        cell.explicit_suffix = override.approved_suffix.upper()
    if override.approved_item_number:
        cell.explicit_item_number = override.approved_item_number.upper()
        match = ITEM_NUMBER_RE.match(cell.explicit_item_number)
        if match:
            cell.explicit_module = match.group("module")
            cell.explicit_suffix = match.group("suffix")
            cell.explicit_sequence_number = int(match.group("number"))


def infer_module_and_suffix(cell: ParsedCell, parent_module: str | None) -> None:
    if cell.explicit_item_number:
        match = ITEM_NUMBER_RE.match(cell.explicit_item_number)
        if match:
            cell.inferred_module = match.group("module")
            cell.inferred_suffix = match.group("suffix")
        return

    if cell.explicit_module:
        cell.inferred_module = cell.explicit_module
    elif cell.syntax_kind == "type_only":
        if parent_module:
            cell.inferred_module = parent_module
        else:
            cell.is_ambiguous = True
            cell.blocker_reason = "ambiguous_module"
            cell.resolution_status = "needs_review"
            cell.action = "review"
            return
    elif cell.syntax_kind == "plain" and parent_module:
        cell.inferred_module = parent_module

    if cell.explicit_suffix:
        cell.inferred_suffix = cell.explicit_suffix
        return

    if cell.syntax_kind in {"module_prefix", "plain"}:
        if not cell.inferred_module:
            if cell.syntax_kind == "plain":
                cell.is_ambiguous = True
                cell.blocker_reason = "ambiguous_module"
                cell.resolution_status = "needs_review"
                cell.action = "review"
            return
        if cell.inferred_module == "UN":
            cell.inferred_suffix = "A" if cell.has_children else "P"
        else:
            cell.inferred_suffix = "A" if cell.has_children else "P"


def parse_cell(
    raw_text: str,
    row_index: int,
    tree_level: int,
    source_column: str,
    parent_raw_text: str,
    parent_module: str | None,
    has_children: bool,
    module_codes: set[str],
    review_override: ReviewOverride | None = None,
) -> ParsedCell:
    cleaned_text = collapse_whitespace(raw_text)
    item_text_without_comment, _, comment_text = cleaned_text.partition("//")
    item_text_without_comment = collapse_whitespace(item_text_without_comment)
    comment = collapse_whitespace(comment_text)

    blocker_reason = None
    quantity = 1
    malformed_quantity = False
    quantity_match = QUANTITY_RE.search(item_text_without_comment)
    if quantity_match:
        quantity = int(quantity_match.group(1))
        base_text = collapse_whitespace(item_text_without_comment[: quantity_match.start()])
    else:
        base_text = item_text_without_comment
        hash_index = base_text.rfind("#")
        if hash_index != -1:
            malformed_quantity = True

    syntax_kind, explicit_item_number, explicit_module, explicit_sequence_number, explicit_suffix, name_portion = _parse_explicit_structure(
        base_text, module_codes
    )
    normalized_name = normalize_item_name(name_portion)
    cell = ParsedCell(
        row_index=row_index,
        tree_level=tree_level,
        raw_text=str(raw_text),
        cleaned_text=cleaned_text,
        item_text_without_comment=item_text_without_comment,
        comment=comment,
        quantity=quantity,
        explicit_item_number=explicit_item_number,
        explicit_module=explicit_module,
        explicit_sequence_number=explicit_sequence_number,
        explicit_suffix=explicit_suffix,
        inferred_module=None,
        inferred_suffix=None,
        normalized_item_name=normalized_name,
        parent_raw_text=parent_raw_text,
        parent_resolved_item_number=None,
        resolution_status="skipped",
        resolved_item_number=None,
        resolved_item_name=None,
        action="skip",
        is_ambiguous=False,
        blocker_reason=blocker_reason,
        syntax_kind=syntax_kind,
        has_children=has_children,
        parent_key=None,
        cell_key=f"r{row_index}_l{tree_level}",
        source_column=source_column,
        review_override=review_override,
    )

    _apply_review_override(cell, review_override)
    infer_module_and_suffix(cell, parent_module)

    if malformed_quantity and cell.blocker_reason is None and not (review_override and review_override.approved_quantity is not None):
        cell.is_ambiguous = True
        cell.blocker_reason = "malformed_quantity"
        cell.resolution_status = "needs_review"
        cell.action = "review"

    return cell


def resolve_candidate(cell: ParsedCell, authority: InventoryAuthority) -> None:
    if cell.blocker_reason in {"ambiguous_module", "ambiguous_type"}:
        cell.resolution_status = "needs_review"
        cell.action = "review"
        return

    forced_create_new = bool(
        cell.review_override and cell.review_override.review_decision == "create_new"
    )
    forced_match_existing = bool(
        cell.review_override and cell.review_override.review_decision == "match_existing"
    )
    forced_skip = bool(cell.review_override and cell.review_override.review_decision == "skip")
    forced_conflict = bool(cell.review_override and cell.review_override.review_decision == "mark_conflict")

    if forced_skip:
        cell.resolution_status = "skipped"
        cell.action = "skip"
        return
    if forced_conflict:
        cell.resolution_status = "conflict"
        cell.blocker_reason = "conflict"
        cell.action = "review"
        return

    if cell.explicit_item_number:
        existing = authority.lookup_by_number(cell.explicit_item_number)
        if existing:
            if cell.normalized_item_name and is_material_name_difference(cell.normalized_item_name, existing.partname):
                if forced_create_new:
                    cell.resolution_status = "new_number_assigned"
                    cell.resolved_item_name = cell.normalized_item_name
                    cell.action = "create_item"
                    cell.resolved_item_number = None
                    return
                if forced_match_existing:
                    cell.resolved_item_number = existing.partnumber
                    cell.resolved_item_name = existing.partname
                    cell.resolution_status = "matched_by_number"
                    cell.action = "use_existing"
                    return
                cell.resolution_status = "conflict"
                cell.blocker_reason = "conflict"
                cell.action = "review"
                return
            cell.resolved_item_number = existing.partnumber
            cell.resolved_item_name = existing.partname
            cell.resolution_status = "matched_by_number"
            cell.action = "use_existing"
            return

        if cell.review_override and cell.review_override.approved_item_number and ITEM_NUMBER_RE.match(cell.review_override.approved_item_number):
            cell.resolved_item_number = cell.review_override.approved_item_number
            cell.resolved_item_name = cell.normalized_item_name
            cell.resolution_status = "new_number_assigned"
            cell.action = "create_item"
            return

        cell.resolution_status = "needs_review"
        cell.blocker_reason = "needs_review"
        cell.action = "review"
        return

    module_code = cell.explicit_module or cell.inferred_module
    suffix = cell.explicit_suffix or cell.inferred_suffix
    matches = authority.lookup_by_name(cell.normalized_item_name)
    review_made_explicit = bool(
        cell.review_override
        and (
            cell.review_override.approved_module
            or cell.review_override.approved_suffix
            or cell.review_override.approved_item_number
        )
    )

    if module_code and suffix:
        exact_matches = authority.lookup_exact(cell.normalized_item_name, module_code, suffix)
        if len(exact_matches) == 1:
            item = exact_matches[0]
            cell.resolved_item_number = item.partnumber
            cell.resolved_item_name = item.partname
            cell.resolution_status = "matched_by_name"
            cell.action = "use_existing"
            return
        if len(exact_matches) > 1:
            cell.is_ambiguous = True
            cell.resolution_status = "needs_review"
            cell.blocker_reason = "needs_review"
            cell.action = "review"
            return

        if forced_match_existing and len(matches) == 1:
            item = matches[0]
            cell.resolved_item_number = item.partnumber
            cell.resolved_item_name = item.partname
            cell.resolution_status = "matched_by_name"
            cell.action = "use_existing"
            return

        if forced_create_new or review_made_explicit or len(exact_matches) == 0:
            cell.resolution_status = "new_number_assigned"
            cell.resolved_item_name = cell.normalized_item_name
            cell.action = "create_item"
            return

    if len(matches) > 1:
        cell.is_ambiguous = True
        cell.resolution_status = "needs_review"
        cell.blocker_reason = "needs_review"
        cell.action = "review"
        return

    cell.resolution_status = "needs_review"
    cell.blocker_reason = cell.blocker_reason or "needs_review"
    cell.action = "review"


def allocate_new_numbers(cells: list[ParsedCell], authority: InventoryAuthority) -> None:
    allocated_by_key: dict[tuple[str, str, str], str] = {}
    allocated_by_explicit_number: dict[str, tuple[str, str, str]] = {}
    used_numbers = set(authority.by_number)
    next_numbers = {module: authority.max_sequence_for_module(module) + 1 for module in authority.module_codes}

    for cell in cells:
        if cell.resolution_status != "new_number_assigned":
            continue
        if cell.resolved_item_number and cell.resolved_item_number in authority.by_number:
            continue

        module_code = cell.explicit_module or cell.inferred_module
        suffix = cell.explicit_suffix or cell.inferred_suffix
        if not module_code or not suffix:
            cell.resolution_status = "needs_review"
            cell.blocker_reason = "ambiguous_module" if not module_code else "ambiguous_type"
            cell.action = "review"
            continue

        key = (module_code, suffix, cell.normalized_item_name)
        if cell.resolved_item_number:
            explicit_number = cell.resolved_item_number
            prior_key = allocated_by_explicit_number.get(explicit_number)
            if prior_key and prior_key != key:
                cell.resolution_status = "conflict"
                cell.blocker_reason = "conflict"
                cell.action = "review"
                continue
            if explicit_number in used_numbers and authority.lookup_by_number(explicit_number) is None:
                cell.resolution_status = "conflict"
                cell.blocker_reason = "conflict"
                cell.action = "review"
                continue
            allocated_by_key.setdefault(key, explicit_number)
            allocated_by_explicit_number[explicit_number] = key
            used_numbers.add(explicit_number)
            cell.resolved_item_number = explicit_number
            cell.resolved_item_name = cell.normalized_item_name
            continue

        if key not in allocated_by_key:
            next_number = next_numbers.setdefault(module_code, authority.max_sequence_for_module(module_code) + 1)
            partnumber = f"{module_code}{str(next_number).zfill(3)}{suffix}"
            while partnumber in used_numbers:
                next_number += 1
                partnumber = f"{module_code}{str(next_number).zfill(3)}{suffix}"
            allocated_by_key[key] = partnumber
            next_numbers[module_code] = next_number + 1
            used_numbers.add(partnumber)

        cell.resolved_item_number = allocated_by_key[key]
        cell.resolved_item_name = cell.normalized_item_name


def annotate_parent_links(cells: list[ParsedCell], cell_by_key: dict[str, ParsedCell]) -> None:
    for cell in cells:
        if not cell.parent_key:
            continue
        parent = cell_by_key[cell.parent_key]
        cell.parent_resolved_item_number = parent.resolved_item_number
        if parent.resolved_item_number is None and cell.resolution_status not in BLOCKER_STATUSES:
            cell.resolution_status = "needs_review"
            cell.blocker_reason = "missing_parent"
            cell.action = "review"


def _build_cell_lookup(cells: list[ParsedCell]) -> dict[str, ParsedCell]:
    return {cell.cell_key: cell for cell in cells}


def harmonize_group_suffixes(cells: list[ParsedCell]) -> None:
    grouped: dict[tuple[str, str], list[ParsedCell]] = {}
    for cell in cells:
        if cell.explicit_item_number or cell.explicit_suffix:
            continue
        module_code = cell.explicit_module or cell.inferred_module
        if not module_code or not cell.normalized_item_name:
            continue
        grouped.setdefault((cell.normalized_item_name, module_code), []).append(cell)

    for _, group in grouped.items():
        if not any(cell.has_children for cell in group):
            continue
        for cell in group:
            cell.inferred_suffix = "A"


def build_tree_path(cell: ParsedCell, cell_lookup: dict[str, ParsedCell]) -> str:
    chain = [cell.normalized_item_name or cell.cleaned_text]
    current = cell
    while current.parent_key:
        parent = cell_lookup[current.parent_key]
        chain.append(parent.resolved_item_name or parent.normalized_item_name or parent.cleaned_text)
        current = parent
    chain.reverse()
    return " > ".join(chain)


def _candidate_display_items(candidates: list[InventoryItem], limit: int = 8) -> str:
    if not candidates:
        return ""
    parts = [f"{item.partnumber}:{item.partname}" for item in candidates[:limit]]
    if len(candidates) > limit:
        parts.append(f"... +{len(candidates) - limit} more")
    return " | ".join(parts)


def _suggested_decision(cell: ParsedCell) -> str:
    if cell.resolution_status == "conflict" or cell.blocker_reason == "conflict":
        return "mark_conflict"
    if (cell.explicit_module or cell.inferred_module) and (cell.explicit_suffix or cell.inferred_suffix):
        return "create_new"
    if cell.blocker_reason == "missing_parent":
        return "resolve_parent_first"
    return "match_existing"


def review_key_for_cell(cell: ParsedCell) -> str:
    module_code = cell.explicit_module or cell.inferred_module or ""
    suffix = cell.explicit_suffix or cell.inferred_suffix or ""
    if cell.explicit_item_number:
        return f"numbered|{cell.explicit_item_number}|{cell.normalized_item_name}"
    return f"name|{cell.normalized_item_name}|{module_code}|{suffix}|{cell.syntax_kind}"


def build_parsed_cells(
    df_mindmap: pd.DataFrame,
    authority: InventoryAuthority,
    overrides: dict[str, ReviewOverride] | None = None,
) -> list[ParsedCell]:
    overrides = overrides or {}
    cells: list[ParsedCell] = []
    cell_by_key: dict[str, ParsedCell] = {}

    for row_zero_based, (_, row) in enumerate(df_mindmap.iterrows(), start=1):
        row_cells = []
        for col_index, column in enumerate(df_mindmap.columns):
            value = row[column]
            if not is_nonempty(value):
                continue
            row_cells.append((col_index, column, value))

        parent_cell: ParsedCell | None = None
        for idx, (col_index, column, value) in enumerate(row_cells):
            has_children = idx < len(row_cells) - 1
            parent_raw_text = parent_cell.raw_text if parent_cell else ""
            parent_module = None
            if parent_cell:
                parent_module = parent_cell.explicit_module or parent_cell.inferred_module

            cell_key = f"r{row_zero_based}_l{col_index}"
            cell = parse_cell(
                raw_text=str(value),
                row_index=row_zero_based,
                tree_level=col_index,
                source_column=str(column),
                parent_raw_text=parent_raw_text,
                parent_module=parent_module,
                has_children=has_children,
                module_codes=authority.module_codes,
                review_override=overrides.get(cell_key),
            )
            if parent_cell:
                cell.parent_key = parent_cell.cell_key
            cells.append(cell)
            cell_by_key[cell.cell_key] = cell
            parent_cell = cell

    harmonize_group_suffixes(cells)
    for cell in cells:
        resolve_candidate(cell, authority)

    allocate_new_numbers(cells, authority)
    annotate_parent_links(cells, cell_by_key)
    return cells


def build_review_csv(cells: list[ParsedCell], authority: InventoryAuthority | None = None) -> pd.DataFrame:
    authority = authority or InventoryAuthority([], pd.DataFrame())
    cell_lookup = _build_cell_lookup(cells)
    duplicate_counts: dict[tuple[str, str | None, str | None], int] = {}
    for cell in cells:
        key = (
            cell.normalized_item_name,
            cell.explicit_module or cell.inferred_module,
            cell.explicit_suffix or cell.inferred_suffix,
        )
        duplicate_counts[key] = duplicate_counts.get(key, 0) + 1

    rows_by_review_key: dict[str, dict[str, object]] = {}
    for cell in cells:
        if cell.resolution_status not in BLOCKER_STATUSES and cell.blocker_reason != "malformed_quantity":
            continue
        if _suggested_decision(cell) == "create_new" and cell.blocker_reason != "malformed_quantity":
            continue
        if cell.blocker_reason == "missing_parent":
            continue
        match_candidates = authority.lookup_by_name(cell.normalized_item_name)
        duplicate_key = (
            cell.normalized_item_name,
            cell.explicit_module or cell.inferred_module,
            cell.explicit_suffix or cell.inferred_suffix,
        )
        review_key = review_key_for_cell(cell)
        if review_key not in rows_by_review_key:
            rows_by_review_key[review_key] = {
                "review_key": review_key,
                "cell_key": cell.cell_key,
                "row": cell.row_index,
                "level": cell.tree_level,
                "tree_path": build_tree_path(cell, cell_lookup),
                "raw_cell": cell.raw_text,
                "cleaned_cell": cell.cleaned_text,
                "normalized_item_name": cell.normalized_item_name,
                "comment": cell.comment,
                "quantity": cell.quantity,
                "parsed_item_number": cell.explicit_item_number,
                "parsed_module": cell.explicit_module,
                "parsed_suffix": cell.explicit_suffix,
                "inferred_module": cell.inferred_module,
                "inferred_suffix": cell.inferred_suffix,
                "resolved_item_number": cell.resolved_item_number,
                "resolved_item_name": cell.resolved_item_name,
                "parent_raw_cell": cell.parent_raw_text,
                "parent_resolved_item_number": cell.parent_resolved_item_number,
                "resolution_status": cell.resolution_status,
                "action": cell.action,
                "is_ambiguous": cell.is_ambiguous,
                "blocker_reason": cell.blocker_reason,
                "occurrence_count": duplicate_counts[duplicate_key],
                "inventory_match_candidates": _candidate_display_items(match_candidates),
                "suggested_decision": _suggested_decision(cell),
                "review_decision": "",
                "approved_module": "",
                "approved_suffix": "",
                "approved_item_name": "",
                "approved_quantity": "",
                "approved_comment": "",
                "approved_item_number": "",
            }
        else:
            rows_by_review_key[review_key]["occurrence_count"] = max(
                int(rows_by_review_key[review_key]["occurrence_count"]),
                duplicate_counts[duplicate_key],
            )
    return pd.DataFrame(rows_by_review_key.values())


def review_workbook_path(review_csv_path: Path) -> Path:
    return review_csv_path.with_suffix(".xlsx")


def _autosize_worksheet(worksheet) -> None:
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 48)


def write_review_workbook(review_df: pd.DataFrame, output_path: Path) -> Path:
    instructions = pd.DataFrame(
        [
            {
                "step": 1,
                "instruction": "Work in the 'review_items' sheet. Filter by blocker_reason or suggested_decision to review items one by one.",
            },
            {
                "step": 2,
                "instruction": "Choose a review_decision: match_existing, create_new, skip, mark_conflict, or resolve_parent_first.",
            },
            {
                "step": 3,
                "instruction": "For match_existing, fill approved_item_number. For create_new, fill approved_module, approved_suffix, and approved_item_name.",
            },
            {
                "step": 4,
                "instruction": "Use approved_quantity only when the source has a malformed quantity marker that needs human correction.",
            },
            {
                "step": 5,
                "instruction": "You may edit approved_comment if the displayed comment should be carried into the final Miro CSV.",
            },
            {
                "step": 6,
                "instruction": "Save the workbook and pass it back with --review. Export/apply stays blocked until every blocker row is resolved.",
            },
        ]
    )
    decision_options = pd.DataFrame(
        {
            "review_decision": ["match_existing", "create_new", "skip", "mark_conflict", "resolve_parent_first"],
            "approved_suffix": ["P", "A", "", "", ""],
        }
    )

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        review_df.to_excel(writer, sheet_name="review_items", index=False)
        instructions.to_excel(writer, sheet_name="instructions", index=False)
        decision_options.to_excel(writer, sheet_name="allowed_values", index=False)

    workbook = load_workbook(output_path)
    review_sheet = workbook["review_items"]
    instructions_sheet = workbook["instructions"]
    allowed_sheet = workbook["allowed_values"]

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in review_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    review_sheet.freeze_panes = "A2"
    review_sheet.auto_filter.ref = review_sheet.dimensions
    instructions_sheet.freeze_panes = "A2"

    decision_col = None
    suffix_col = None
    for index, header in enumerate(review_df.columns, start=1):
        if header == "review_decision":
            decision_col = get_column_letter(index)
        if header == "approved_suffix":
            suffix_col = get_column_letter(index)

    if decision_col:
        validation = DataValidation(
            type="list",
            formula1="=allowed_values!$A$2:$A$6",
            allow_blank=True,
        )
        review_sheet.add_data_validation(validation)
        validation.add(f"{decision_col}2:{decision_col}{review_sheet.max_row}")

    if suffix_col:
        validation = DataValidation(
            type="list",
            formula1="=allowed_values!$B$2:$B$3",
            allow_blank=True,
        )
        review_sheet.add_data_validation(validation)
        validation.add(f"{suffix_col}2:{suffix_col}{review_sheet.max_row}")

    _autosize_worksheet(review_sheet)
    _autosize_worksheet(instructions_sheet)
    _autosize_worksheet(allowed_sheet)
    workbook.save(output_path)
    return output_path


def read_review_overrides(path: Path) -> dict[str, ReviewOverride]:
    if path.suffix.lower() == ".xlsx":
        df = pd.read_excel(path, sheet_name="review_items", dtype=str).fillna("")
    else:
        df = pd.read_csv(path, dtype=str).fillna("")
    overrides: dict[str, ReviewOverride] = {}
    for _, row in df.iterrows():
        review_key = collapse_whitespace(row.get("review_key", ""))
        cell_key = collapse_whitespace(row.get("cell_key", ""))
        override_key = review_key or cell_key
        if not override_key:
            continue
        override = ReviewOverride(
            review_decision=collapse_whitespace(row.get("review_decision", "")) or None,
            approved_module=collapse_whitespace(row.get("approved_module", "")) or None,
            approved_suffix=collapse_whitespace(row.get("approved_suffix", "")) or None,
            approved_item_name=collapse_whitespace(row.get("approved_item_name", "")) or None,
            approved_quantity=int(collapse_whitespace(row.get("approved_quantity", ""))) if collapse_whitespace(row.get("approved_quantity", "")) else None,
            approved_comment=collapse_whitespace(row.get("approved_comment", "")) or None,
            approved_item_number=collapse_whitespace(row.get("approved_item_number", "")).upper() or None,
        )
        if any(
            (
                override.approved_module,
                override.approved_suffix,
                override.approved_item_name,
                override.approved_quantity,
                override.approved_comment,
                override.approved_item_number,
                override.review_decision,
            )
        ):
            overrides[override_key] = override
    return overrides


def apply_review_csv(
    df_mindmap: pd.DataFrame,
    authority: InventoryAuthority,
    review_path: Path,
) -> list[ParsedCell]:
    workbook_overrides = read_review_overrides(review_path)
    initial_cells = build_parsed_cells(df_mindmap, authority)
    per_cell_overrides: dict[str, ReviewOverride] = {}
    for cell in initial_cells:
        review_key = review_key_for_cell(cell)
        override = workbook_overrides.get(review_key) or workbook_overrides.get(cell.cell_key)
        if override:
            per_cell_overrides[cell.cell_key] = override
    return build_parsed_cells(df_mindmap, authority, overrides=per_cell_overrides)


def format_miro_cell(cell: ParsedCell) -> str:
    if not cell.resolved_item_number or not cell.resolved_item_name:
        raise ValueError(f"Cell {cell.cell_key} is unresolved and cannot be exported.")

    rendered = f"{cell.resolved_item_number}:{cell.resolved_item_name}"
    if cell.quantity > 1:
        rendered += f" #{cell.quantity}"
    if cell.comment:
        rendered += f" //{cell.comment}"
    return rendered


def dataframe_to_opml_rows(df_mindmap: pd.DataFrame, cells: list[ParsedCell]) -> list[list[str]]:
    cell_lookup = {(cell.row_index, cell.tree_level): cell for cell in cells}
    rows: list[list[str]] = []
    for row_index in range(1, len(df_mindmap) + 1):
        path_parts: list[str] = []
        for col_index, _column in enumerate(df_mindmap.columns):
            if not is_nonempty(df_mindmap.iloc[row_index - 1, col_index]):
                continue
            path_parts.append(format_miro_cell(cell_lookup[(row_index, col_index)]))
        if path_parts:
            rows.append(path_parts)
    return rows


def write_miro_opml(df_mindmap: pd.DataFrame, cells: list[ParsedCell], output_path: Path, title: str) -> Path:
    rows = dataframe_to_opml_rows(df_mindmap, cells)
    if not rows:
        raise ValueError("No rows available to export to OPML.")

    root_text = rows[0][0]
    opml = ET.Element("opml", {"version": "2.0"})
    head = ET.SubElement(opml, "head")
    ET.SubElement(head, "title").text = title
    body = ET.SubElement(opml, "body")
    root_outline = ET.SubElement(body, "outline", {"text": root_text})

    previous_path = [root_text]
    previous_nodes = [root_outline]

    for row in rows[1:]:
        if row[0] != root_text:
            raise ValueError("OPML export requires a single root outline.")

        common_depth = 0
        max_common = min(len(previous_path), len(row))
        while common_depth < max_common and previous_path[common_depth] == row[common_depth]:
            common_depth += 1

        current_nodes = previous_nodes[:common_depth]
        for index in range(common_depth, len(row)):
            parent = body if index == 0 else current_nodes[index - 1]
            new_node = ET.SubElement(parent, "outline", {"text": row[index]})
            if index == 0:
                root_outline = new_node
            if len(current_nodes) > index:
                current_nodes[index] = new_node
            else:
                current_nodes.append(new_node)

        previous_path = row
        previous_nodes = current_nodes

    xml_lines = ['<?xml version="1.0" encoding="UTF-8"?>', '<opml version="2.0">', "  <head>", f"    <title>{title}</title>", "  </head>", "  <body>"]

    def emit_outline(node: ET.Element, indent: int) -> None:
        spaces = "  " * indent
        text = node.attrib["text"]
        children = [child for child in node if child.tag == "outline"]
        if not children:
            xml_lines.append(f'{spaces}<outline text="{_xml_escape(text)}"/>')
            return
        xml_lines.append(f'{spaces}<outline text="{_xml_escape(text)}">')
        for child in children:
            emit_outline(child, indent + 1)
        xml_lines.append(f"{spaces}</outline>")

    emit_outline(root_outline, 2)
    xml_lines.extend(["  </body>", "</opml>"])
    output_path.write_text("\n".join(xml_lines) + "\n", encoding="utf-8")
    return output_path


def _xml_escape(value: str) -> str:
    return (
        value.replace("&", "&amp;")
        .replace('"', "&quot;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


def write_miro_csv(df_mindmap: pd.DataFrame, cells: list[ParsedCell], output_path: Path) -> Path:
    cell_lookup = {(cell.row_index, cell.tree_level): cell for cell in cells}
    export_df = df_mindmap.copy()

    for row_index in range(1, len(export_df) + 1):
        for col_index, column in enumerate(export_df.columns):
            if not is_nonempty(export_df.iloc[row_index - 1, col_index]):
                continue
            cell = cell_lookup[(row_index, col_index)]
            export_df.iloc[row_index - 1, col_index] = format_miro_cell(cell)

    export_df.to_csv(output_path, index=False)
    return output_path


def write_miro_output(
    df_mindmap: pd.DataFrame,
    cells: list[ParsedCell],
    output_path: Path,
    input_format: str,
    title: str,
) -> Path:
    if input_format == "opml":
        return write_miro_opml(df_mindmap, cells, output_path, title=title)
    return write_miro_csv(df_mindmap, cells, output_path)


def update_inventory_workbook(inventory_path: Path, cells: list[ParsedCell]) -> int:
    df_inventory = pd.read_excel(inventory_path)
    existing_columns = list(df_inventory.columns)
    if "partnumber" not in df_inventory.columns:
        df_inventory["partnumber"] = pd.NA
    if "partname" not in df_inventory.columns:
        df_inventory["partname"] = pd.NA

    existing_numbers = set(df_inventory["partnumber"].astype(str))
    new_rows: list[dict[str, object]] = []
    seen_new_numbers: set[str] = set()

    for cell in cells:
        if cell.resolution_status != "new_number_assigned":
            continue
        if not cell.resolved_item_number or not cell.resolved_item_name:
            continue
        if cell.resolved_item_number in existing_numbers or cell.resolved_item_number in seen_new_numbers:
            continue
        row = {column: pd.NA for column in df_inventory.columns}
        row["partnumber"] = cell.resolved_item_number
        row["partname"] = cell.resolved_item_name
        new_rows.append(row)
        seen_new_numbers.add(cell.resolved_item_number)

    if not new_rows:
        return 0

    updated_inventory = pd.concat([df_inventory, pd.DataFrame(new_rows)], ignore_index=True)
    updated_inventory = updated_inventory.reindex(columns=existing_columns)
    updated_inventory.to_excel(inventory_path, index=False)
    return len(new_rows)


def blocks_miro_export(cell: ParsedCell) -> bool:
    if cell.resolution_status == "conflict":
        return True
    if not cell.resolved_item_number:
        return True
    if cell.blocker_reason in {"ambiguous_module", "ambiguous_type", "needs_review", "malformed_quantity"}:
        return True
    return False


def _table_columns(connection: sqlite3.Connection, table_name: str) -> list[str]:
    result = connection.execute(f"PRAGMA table_info({table_name})").fetchall()
    return [row[1] for row in result]


def _first_present(columns: list[str], candidates: list[str]) -> str | None:
    for candidate in candidates:
        if candidate in columns:
            return candidate
    return None


def write_database_records(database_path: Path, cells: list[ParsedCell]) -> None:
    blockers = [cell for cell in cells if cell.resolution_status in BLOCKER_STATUSES]
    if blockers:
        raise RuntimeError("Refusing database write because blocker rows still exist.")

    connection = sqlite3.connect(database_path)
    try:
        tables = {row[0] for row in connection.execute("SELECT name FROM sqlite_master WHERE type='table'")}
        required = {"items", "item_revisions", "bom_links"}
        missing = required - tables
        if missing:
            raise RuntimeError(f"Database is missing required tables: {', '.join(sorted(missing))}")

        item_columns = _table_columns(connection, "items")
        revision_columns = _table_columns(connection, "item_revisions")
        link_columns = _table_columns(connection, "bom_links")

        item_number_col = _first_present(item_columns, ["partnumber", "item_number"])
        item_name_col = _first_present(item_columns, ["partname", "name", "item_name"])
        item_module_col = _first_present(item_columns, ["module_code", "module"])
        item_type_col = _first_present(item_columns, ["item_type", "type"])
        if not item_number_col or not item_name_col:
            raise RuntimeError("Unsupported items table schema.")

        revision_item_fk = _first_present(revision_columns, ["item_id"])
        revision_name_col = _first_present(revision_columns, ["name", "item_name"])
        revision_comment_col = _first_present(revision_columns, ["comments", "comment"])
        revision_default_col = _first_present(revision_columns, ["is_default"])
        if not revision_item_fk:
            raise RuntimeError("Unsupported item_revisions table schema.")

        link_parent_col = _first_present(link_columns, ["parent_revision_id", "parent_item_revision_id"])
        link_child_col = _first_present(link_columns, ["child_revision_id", "child_item_revision_id"])
        link_qty_col = _first_present(link_columns, ["quantity"])
        link_comment_col = _first_present(link_columns, ["comments", "comment"])
        if not link_parent_col or not link_child_col or not link_qty_col:
            raise RuntimeError("Unsupported bom_links table schema.")

        item_revision_id_by_number: dict[str, int] = {}
        item_id_by_number: dict[str, int] = {}

        def ensure_item(cell: ParsedCell) -> tuple[int, int]:
            assert cell.resolved_item_number is not None
            assert cell.resolved_item_name is not None

            if cell.resolved_item_number in item_revision_id_by_number:
                return item_id_by_number[cell.resolved_item_number], item_revision_id_by_number[cell.resolved_item_number]

            existing_row = connection.execute(
                f"SELECT id FROM items WHERE {item_number_col} = ?",
                (cell.resolved_item_number,),
            ).fetchone()

            module_code = (cell.explicit_module or cell.inferred_module or "")[:3]
            item_type = "assembly" if (cell.explicit_suffix or cell.inferred_suffix) == "A" else "part"

            if existing_row:
                item_id = int(existing_row[0])
            else:
                insert_columns = [item_number_col, item_name_col]
                insert_values = [cell.resolved_item_number, cell.resolved_item_name]
                if item_module_col:
                    insert_columns.append(item_module_col)
                    insert_values.append(module_code)
                if item_type_col:
                    insert_columns.append(item_type_col)
                    insert_values.append(item_type)
                placeholders = ", ".join("?" for _ in insert_columns)
                connection.execute(
                    f"INSERT INTO items ({', '.join(insert_columns)}) VALUES ({placeholders})",
                    insert_values,
                )
                item_id = int(connection.execute("SELECT last_insert_rowid()").fetchone()[0])

            revision_row = connection.execute(
                f"SELECT id FROM item_revisions WHERE {revision_item_fk} = ? ORDER BY id LIMIT 1",
                (item_id,),
            ).fetchone()

            if revision_row:
                revision_id = int(revision_row[0])
            else:
                insert_columns = [revision_item_fk]
                insert_values = [item_id]
                if revision_name_col:
                    insert_columns.append(revision_name_col)
                    insert_values.append(cell.resolved_item_name)
                if revision_comment_col:
                    insert_columns.append(revision_comment_col)
                    insert_values.append("")
                if revision_default_col:
                    insert_columns.append(revision_default_col)
                    insert_values.append(1)
                placeholders = ", ".join("?" for _ in insert_columns)
                connection.execute(
                    f"INSERT INTO item_revisions ({', '.join(insert_columns)}) VALUES ({placeholders})",
                    insert_values,
                )
                revision_id = int(connection.execute("SELECT last_insert_rowid()").fetchone()[0])

            item_id_by_number[cell.resolved_item_number] = item_id
            item_revision_id_by_number[cell.resolved_item_number] = revision_id
            return item_id, revision_id

        unique_cells = {}
        for cell in cells:
            if cell.resolved_item_number:
                unique_cells.setdefault(cell.resolved_item_number, cell)
        for cell in unique_cells.values():
            ensure_item(cell)

        seen_links: set[tuple[int, int, int, str]] = set()
        for cell in cells:
            if not cell.parent_resolved_item_number or not cell.resolved_item_number:
                continue
            if cell.parent_resolved_item_number == cell.resolved_item_number:
                continue
            _, parent_revision_id = ensure_item(unique_cells[cell.parent_resolved_item_number])
            _, child_revision_id = ensure_item(unique_cells[cell.resolved_item_number])
            link_key = (parent_revision_id, child_revision_id, cell.quantity, cell.comment)
            if link_key in seen_links:
                continue
            seen_links.add(link_key)

            where_parts = [
                f"{link_parent_col} = ?",
                f"{link_child_col} = ?",
                f"{link_qty_col} = ?",
            ]
            where_values: list[object] = [parent_revision_id, child_revision_id, cell.quantity]
            if link_comment_col:
                where_parts.append(f"COALESCE({link_comment_col}, '') = ?")
                where_values.append(cell.comment)
            existing_link = connection.execute(
                f"SELECT id FROM bom_links WHERE {' AND '.join(where_parts)}",
                where_values,
            ).fetchone()
            if existing_link:
                continue

            insert_columns = [link_parent_col, link_child_col, link_qty_col]
            insert_values: list[object] = [parent_revision_id, child_revision_id, cell.quantity]
            if link_comment_col:
                insert_columns.append(link_comment_col)
                insert_values.append(cell.comment)
            placeholders = ", ".join("?" for _ in insert_columns)
            connection.execute(
                f"INSERT INTO bom_links ({', '.join(insert_columns)}) VALUES ({placeholders})",
                insert_values,
            )

        connection.commit()
    finally:
        connection.close()


def infer_subsystem(input_path: Path, fallback: str | None = None) -> str:
    if fallback:
        return fallback.upper()
    stem = input_path.stem.upper()
    for token in re.split(r"[\s_\-]+", stem):
        if token in DEFAULT_MODULE_CODES:
            return token
    return "BOM"


def review_output_path(repo_root: Path, subsystem: str) -> Path:
    review_dir = repo_root / "review"
    review_dir.mkdir(parents=True, exist_ok=True)
    return review_dir / f"bom_import_review_{subsystem}.csv"


def miro_output_path(repo_root: Path, input_path: Path) -> Path:
    output_dir = repo_root / "output"
    output_dir.mkdir(parents=True, exist_ok=True)
    extension = ".opml" if input_path.suffix.lower() == ".opml" else ".csv"
    return output_dir / f"{input_path.stem}_partnumbered_approved{extension}"


def run_import(
    input_path: Path,
    inventory_path: Path,
    review_path: Path | None = None,
) -> tuple[pd.DataFrame, list[ParsedCell], InventoryAuthority, str]:
    authority = InventoryAuthority.from_inventory_file(inventory_path)
    df_raw, input_format = load_bom_input(input_path)
    df_repaired = df_raw if input_format == "opml" else repair_mindmap_tree(df_raw)
    if review_path:
        cells = apply_review_csv(df_repaired, authority, review_path)
    else:
        cells = build_parsed_cells(df_repaired, authority)
    return df_repaired, cells, authority, input_format


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Dry-run-first BOM CSV digester for Miro/mindmap exports.")
    parser.add_argument("--subsystem", help="Subsystem code used for output naming.", default=None)
    parser.add_argument("--input", required=True, help="Path to the BOM/mindmap CSV export.")
    parser.add_argument("--inventory", help="Path to the current inventory workbook (.xlsx).", default=None)
    parser.add_argument("--review", help="Approved review CSV path for second-pass resolution.", default=None)
    parser.add_argument("--dry-run", action="store_true", help="Write only the review CSV.")
    parser.add_argument("--export-miro", action="store_true", help="Write the final Miro-compatible CSV after blockers are resolved.")
    parser.add_argument("--apply-db", action="store_true", help="Write resolved items and bom_links into a SQLite database.")
    parser.add_argument("--database", help="SQLite database path for --apply-db mode.", default=None)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    repo_root = Path(__file__).resolve().parent
    input_path = Path(args.input)
    inventory_path = Path(args.inventory) if args.inventory else default_inventory_path(repo_root)
    review_path = Path(args.review) if args.review else None
    subsystem = infer_subsystem(input_path, args.subsystem)
    dry_run = args.dry_run or (not args.export_miro and not args.apply_db)

    df_repaired, cells, authority, input_format = run_import(input_path=input_path, inventory_path=inventory_path, review_path=review_path)

    review_df = build_review_csv(cells, authority=authority)
    review_path_out = review_output_path(repo_root, subsystem)
    if dry_run or not review_df.empty:
        review_df.to_csv(review_path_out, index=False)
        workbook_path = write_review_workbook(review_df, review_workbook_path(review_path_out))
        print(f"Review CSV written: {review_path_out}")
        print(f"Review workbook written: {workbook_path}")

    blockers = [cell for cell in cells if cell.resolution_status in BLOCKER_STATUSES]
    if args.export_miro:
        export_blockers = [cell for cell in cells if blocks_miro_export(cell)]
        if export_blockers:
            print(f"Blocked: {len(export_blockers)} unresolved/conflicting rows remain. Review CSV must be resolved before export/apply.")
            return 1

    if blockers and args.apply_db:
        print(f"Blocked: {len(blockers)} unresolved/conflicting rows remain. Review CSV must be resolved before export/apply.")
        return 1

    if args.export_miro:
        added_rows = update_inventory_workbook(inventory_path, cells)
        if added_rows:
            print(f"Inventory workbook updated: {inventory_path} (+{added_rows} items)")
        miro_path = write_miro_output(
            df_repaired,
            cells,
            miro_output_path(repo_root, input_path),
            input_format=input_format,
            title=input_path.name,
        )
        print(f"Miro output written: {miro_path}")

    if args.apply_db:
        if not args.database:
            raise SystemExit("--apply-db requires --database PATH")
        write_database_records(Path(args.database), cells)
        print(f"Database records written: {args.database}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
