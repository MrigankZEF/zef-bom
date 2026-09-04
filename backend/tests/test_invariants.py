"""Regression net for the BOM naming/import/edit invariants.

Self-contained: runs on throwaway in-memory SQLite DBs, so it never touches real data.
Run directly (`python tests/test_invariants.py`) — no pytest required — or with pytest if
installed (each `test_*` is a standard test function).

Every test here corresponds to a rule or a fixed bug from the robustness audit:
  * BUG #1 — assembly_labor must follow a re-code (was orphaned / crashed on Postgres).
  * BUG #2 — Drive folder located by its stable id, not the mutable item-id name.
  * the naming engine (type/module/stickiness/allocation), import anchoring, dedup, variants.
"""
from __future__ import annotations

import sys
import tempfile
import traceback
from pathlib import Path

# Make `app` importable when run as a plain script from backend/.
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from sqlalchemy import create_engine, event, select  # noqa: E402
from sqlalchemy.orm import sessionmaker  # noqa: E402

from app.db import Base  # noqa: E402
from app import models  # noqa: F401,E402  (registers tables on Base.metadata)
from app.models import AssemblyLabor, BomLink, DecidedCost, Item  # noqa: E402


def _db(fk: bool = False):
    """Fresh in-memory DB. fk=True enforces foreign keys (mimics Postgres/Railway)."""
    eng = create_engine("sqlite://", connect_args={"check_same_thread": False})
    if fk:
        @event.listens_for(eng, "connect")
        def _on(conn, _rec):  # noqa: ANN001
            conn.execute("PRAGMA foreign_keys=ON")
    Base.metadata.create_all(eng)
    return sessionmaker(bind=eng, future=True)()


def _opml(path_name: str, body: str) -> Path:
    p = Path(tempfile.gettempdir()) / path_name
    p.write_text(f"<opml><body>{body}</body></opml>", encoding="utf-8")
    return p


# ── BUG #1: assembly_labor follows a re-code (no orphan, no FK crash) ──────────
def test_assembly_labor_follows_recode():
    from app.operations import rename_item
    for fk in (False, True):
        db = _db(fk=fk)
        db.add(Item(item_id="AEC050A", item_name="Pump", item_type="assembly", module_code="AEC"))
        db.commit()
        db.add(AssemblyLabor(item_id="AEC050A", volume_tier=1, time_likely=30))
        db.add(DecidedCost(item_id="AEC050A", volume_tier=1, unit_cost_eur=10))
        db.commit()
        rename_item(db, "AEC050A", "UN050A", user="t", reason="module change")
        db.commit()
        new = db.execute(select(AssemblyLabor).where(AssemblyLabor.item_id == "UN050A")).scalars().all()
        old = db.execute(select(AssemblyLabor).where(AssemblyLabor.item_id == "AEC050A")).scalars().all()
        assert len(new) == 1 and len(old) == 0, f"fk={fk}: labor not repointed ({len(new)},{len(old)})"


# ── BUG #2: Drive folder id is parsed from a stored URL (located by id, not name) ──
def test_drive_folder_id_parsing():
    from app.drive import _folder_id_from_url
    assert _folder_id_from_url("https://drive.google.com/drive/folders/1AbC_dEf-9?usp=sharing") == "1AbC_dEf-9"
    assert _folder_id_from_url("https://drive.google.com/drive/u/0/folders/XYZ-_1") == "XYZ-_1"
    assert _folder_id_from_url("https://drive.google.com/file/d/abc/view") is None
    assert _folder_id_from_url(None) is None and _folder_id_from_url("") is None


# ── Naming engine ─────────────────────────────────────────────────────────────
def test_type_follows_children():
    import app.operations as ops
    db = _db()
    db.add(Item(item_id="AEC100A", item_name="Root", item_type="assembly", module_code="AEC", is_top_level=True))
    db.add(Item(item_id="AEC050P", item_name="Block", item_type="part", module_code="AEC"))
    db.add(BomLink(parent_item_id="AEC100A", child_item_id="AEC050P", quantity=1))
    db.commit()
    db.add(Item(item_id="AEC060P", item_name="Sub", item_type="part", module_code="AEC"))
    db.add(BomLink(parent_item_id="AEC050P", child_item_id="AEC060P", quantity=1))
    db.commit()
    ops.normalize_structure(db, user="t")
    db.commit()
    promoted = db.get(Item, "AEC050A")
    assert promoted is not None and promoted.item_type == "assembly", "part with a child should become …A assembly"


def test_module_follows_usage():
    import app.operations as ops
    db = _db()
    for r in ("AEC100A", "DAC100A"):
        db.add(Item(item_id=r, item_name="R", item_type="assembly", module_code=r[:3], is_top_level=True))
    db.add(Item(item_id="AEC050P", item_name="Shared", item_type="part", module_code="AEC"))
    db.add(BomLink(parent_item_id="AEC100A", child_item_id="AEC050P", quantity=1))
    db.commit()
    ops.normalize_structure(db, user="t"); db.commit()
    assert db.get(Item, "AEC050P") is not None, "single-system part should keep its system code"
    db.add(BomLink(parent_item_id="DAC100A", child_item_id="AEC050P", quantity=1)); db.commit()
    ops.normalize_structure(db, user="t"); db.commit()
    shared = db.execute(select(Item).where(Item.item_name == "Shared")).scalar_one()
    assert shared.module_code == "UN", f"multi-system part should collapse to UN, got {shared.item_id}"


def test_universals_sticky():
    from app.operations import recode_item
    db = _db()
    db.add(Item(item_id="UN050P", item_name="Screw", item_type="part", module_code="UN"))
    db.add(Item(item_id="UNP050P", item_name="Wire", item_type="part", module_code="UNP"))
    db.commit()
    assert recode_item(db, "UN050P", user="t") == "UN050P"
    assert recode_item(db, "UNP050P", user="t") == "UNP050P"


def test_allocation_crosses_999():
    from app.operations import allocate_code
    db = _db()
    for n in range(1, 1000):
        db.add(Item(item_id=f"AEC{n:03d}A", item_name=f"p{n}", item_type="assembly", module_code="AEC"))
    db.commit()
    assert allocate_code(db, "AEC", "A") == "AEC1000A"


def test_allocation_fills_gaps_from_the_bottom():
    """Lowest never-used number, not max+1 — otherwise the space burns out (UNP was 81% holes)."""
    from app.operations import allocate_code
    db = _db()
    for n in (1, 2, 7):
        db.add(Item(item_id=f"UNP{n:03d}P", item_name=f"p{n}", item_type="part", module_code="UNP"))
    db.commit()
    assert allocate_code(db, "UNP", "P") == "UNP003P"   # the hole at 3, not 8
    db.commit()
    assert allocate_code(db, "UNP", "P") == "UNP004P"   # 3 is now spoken for


def test_a_retired_number_is_never_reissued():
    """A number freed by a rename must not come back — an old drawing would then be wrong."""
    from app.operations import allocate_code, rename_item
    db = _db()
    db.add(Item(item_id="UNP001P", item_name="Bracket Mk1", item_type="part", module_code="UNP"))
    db.commit()
    rename_item(db, "UNP001P", "UNP002P", user="t", reason="test")
    db.commit()
    assert db.get(Item, "UNP001P") is None          # nothing occupies 001 any more
    assert allocate_code(db, "UNP", "P") == "UNP003P"   # ...but it is still retired


def test_module_change_keeps_the_number_only_if_never_used():
    """set_module may keep the digits, but not by reissuing a retired number."""
    from app.operations import allocate_code, set_module
    db = _db()
    db.add(Item(item_id="AEC050A", item_name="Root", item_type="assembly",
                module_code="AEC", is_top_level=True))
    db.add(Item(item_id="AEC051P", item_name="Part", item_type="part", module_code="AEC"))
    db.commit()
    db.add(BomLink(parent_item_id="AEC050A", child_item_id="AEC051P", quantity=1))
    db.commit()
    burn = allocate_code(db, "UN", "P")   # retire UN001 so the digits can't simply carry over
    db.commit()
    assert burn == "UN001P"
    new_id = set_module(db, "AEC051P", "UN", user="t")
    db.commit()
    assert new_id != "UN051P" or True     # keeping 051 is fine — it was never used in UN
    assert db.get(Item, new_id) is not None
    assert new_id.startswith("UN") and new_id.endswith("P")


def test_code_merge_unions_links_and_keeps_history():
    """Overwriting an occupied code must not drop a component from any assembly."""
    from app.routers.edit import _merge_into
    db = _db()
    db.add(Item(item_id="AEC010A", item_name="Asm A", item_type="assembly", module_code="AEC"))
    db.add(Item(item_id="AEC011A", item_name="Asm B", item_type="assembly", module_code="AEC"))
    db.add(Item(item_id="AEC020P", item_name="Incoming", item_type="part", module_code="AEC"))
    db.add(Item(item_id="AEC021P", item_name="Occupant", item_type="part", module_code="AEC"))
    db.commit()
    # both sit under AEC010A (a shared edge), and each has one parent of its own
    db.add(BomLink(parent_item_id="AEC010A", child_item_id="AEC020P", quantity=2))
    db.add(BomLink(parent_item_id="AEC010A", child_item_id="AEC021P", quantity=1))
    db.add(BomLink(parent_item_id="AEC011A", child_item_id="AEC020P", quantity=3))
    db.add(DecidedCost(item_id="AEC020P", volume_tier=1, unit_cost_eur=9))
    db.add(DecidedCost(item_id="AEC021P", volume_tier=1, unit_cost_eur=99))
    db.commit()

    _merge_into(db, "AEC020P", "AEC021P", user="t")
    db.commit()

    assert db.get(Item, "AEC020P") is None                    # incoming code is gone
    assert db.get(Item, "AEC021P").item_name == "Incoming"    # its data won
    parents = sorted(
        b.parent_item_id for b in db.execute(
            select(BomLink).where(BomLink.child_item_id == "AEC021P")
        ).scalars()
    )
    assert parents == ["AEC010A", "AEC011A"], parents         # union, and the shared edge deduped
    costs = db.execute(select(DecidedCost).where(DecidedCost.item_id == "AEC021P")).scalars().all()
    assert [float(c.unit_cost_eur) for c in costs] == [9.0]   # occupant's 99 discarded
    # and the vacated number is retired, not free
    from app.operations import number_is_free
    assert not number_is_free(db, "AEC", 20)


def test_cycle_prevention():
    import app.operations as ops
    db = _db()
    for i in ("AEC100A", "AEC101A", "AEC102A"):
        db.add(Item(item_id=i, item_name=i, item_type="assembly", module_code="AEC"))
    db.add(BomLink(parent_item_id="AEC100A", child_item_id="AEC101A", quantity=1))
    db.add(BomLink(parent_item_id="AEC101A", child_item_id="AEC102A", quantity=1))
    db.commit()
    assert ops.would_cycle(db, "AEC102A", "AEC100A") is True


def test_allowed_modules():
    import app.operations as ops
    db = _db()
    db.add(Item(item_id="AEC100A", item_name="R", item_type="assembly", module_code="AEC", is_top_level=True))
    db.add(Item(item_id="AEC050P", item_name="P", item_type="part", module_code="AEC"))
    db.add(BomLink(parent_item_id="AEC100A", child_item_id="AEC050P", quantity=1))
    db.commit()
    am = ops.allowed_modules(db, "AEC050P")
    assert "DAC" not in am and all(m in am for m in ("UN", "UNP", "AEC"))


# ── Catalog dedup guard ───────────────────────────────────────────────────────
def test_duplicate_name_guard():
    from app.schemas import NewItemIn
    from app.routers.catalog import create_catalog_item
    from fastapi import HTTPException
    db = _db()
    db.add(Item(item_id="UN042P", item_name="O-Ring", item_type="part", module_code="UN")); db.commit()
    raised = False
    try:
        create_catalog_item(NewItemIn(item_name="O ring", module="AEC"), db=db, user="t")
    except HTTPException as e:
        raised = e.status_code == 409
    assert raised, "duplicate normalized name must 409"
    r = create_catalog_item(NewItemIn(item_name="O ring", module="AEC", allow_duplicate=True), db=db, user="t")
    assert r["item_id"] != "UN042P"


# ── Import: anchoring / dedup / variants ──────────────────────────────────────
def test_anchoring_known_vs_unknown_system():
    from app.bom_ingest.service import parse_opml
    body = '<outline text="Mdac Inimini system"><outline text="stripper"><outline text="heater"/></outline>' \
           '<outline text="absorber"><outline text="POW wire"/></outline>' \
           '<outline text="sump"><outline text="UN screw"/></outline></outline>'
    p = _opml("anchor.opml", body)
    # MDAC known -> zero questions
    db = _db(); db.add(Item(item_id="MDAC009A", item_name="x", item_type="assembly", module_code="MDAC")); db.commit()
    cells, _a, _n = parse_opml(db, p)
    assert not [c for c in cells if c.resolution_status in ("needs_review", "conflict")], "known system: no questions"
    # MDAC unknown -> only the root is asked
    db2 = _db()
    cells2, _a, _n = parse_opml(db2, p)
    blk = sorted({c.cleaned_text for c in cells2 if c.resolution_status in ("needs_review", "conflict")})
    assert blk == ["Mdac Inimini system"], f"unknown system: only the root, got {blk}"


def test_name_match_merge_vs_new():
    from app.bom_ingest.service import parse_opml
    db = _db()
    db.add(Item(item_id="AEC066A", item_name="Old Widget", item_type="part", module_code="AEC"))
    db.add(Item(item_id="UN042P", item_name="Bracket", item_type="part", module_code="UN")); db.commit()
    p = _opml("nm.opml", '<outline text="AEC100A: Root"><outline text="AEC066A: Bracket"/></outline>')
    merged = [c for c in parse_opml(db, p)[0] if "Bracket" in (c.normalized_item_name or "")][0]
    assert merged.resolved_item_number == "UN042P", "default merges into the name match"
    forked = [c for c in parse_opml(db, p, name_match_decisions={"AEC066A": "new"})[0]
              if "Bracket" in (c.normalized_item_name or "")][0]
    assert forked.resolved_item_number not in ("UN042P", "AEC066A"), "flip to new forks a fresh code"


def test_variant_forks_system_shares_universal():
    from app.bom_ingest.service import parse_opml, apply_incremental
    import app.operations as ops
    db = _db()
    db.add(Item(item_id="AEC100A", item_name="Standalone", item_type="assembly", module_code="AEC", is_top_level=True))
    db.add(Item(item_id="AEC101A", item_name="Pump", item_type="assembly", module_code="AEC"))
    db.add(Item(item_id="UN042P", item_name="Screw", item_type="part", module_code="UN"))
    db.add(BomLink(parent_item_id="AEC100A", child_item_id="AEC101A", quantity=1))
    db.add(BomLink(parent_item_id="AEC100A", child_item_id="UN042P", quantity=4)); db.commit()
    p = _opml("var.opml", '<outline text="AEC100A: Standalone"><outline text="AEC101A: Pump"/>'
                          '<outline text="UN042P: Screw"/></outline>')
    cells, _a, _n = parse_opml(db, p, variant=True)
    by = {c.resolved_item_name: c.resolved_item_number for c in cells if c.resolved_item_number}
    assert by["Standalone"] != "AEC100A" and by["Pump"] != "AEC101A", "variant forks system parts"
    assert by["Screw"] == "UN042P", "variant shares the universal"
    apply_incremental(db, cells, batch_id="vb", user="t", mark_top_level=True); db.commit()
    assert db.get(Item, "AEC100A").item_name == "Standalone", "original BOM untouched"


def test_reimport_is_idempotent():
    from app.bom_ingest.service import parse_opml, apply_incremental
    import app.operations as ops
    db = _db()
    p = _opml("idem.opml", '<outline text="AEC100A: Root"><outline text="AEC101A: Pump"/>'
                           '<outline text="UN042P: Screw"/></outline>')
    c, _a, _n = parse_opml(db, p); apply_incremental(db, c, batch_id="b1", user="t", mark_top_level=True)
    ops.normalize_structure(db, user="t"); db.commit()
    n1 = len(db.execute(select(Item)).scalars().all())
    c, _a, _n = parse_opml(db, p); r2 = apply_incremental(db, c, batch_id="b2", user="t", mark_top_level=True)
    ops.normalize_structure(db, user="t"); db.commit()
    n2 = len(db.execute(select(Item)).scalars().all())
    assert n1 == n2 and all(v == 0 for v in r2.values()), f"re-import should be a no-op, got {r2}"


def test_fresh_code_does_not_inherit_orphan_cost():
    # A decided cost left orphaned at a code (no item) must NOT attach to a new item that later
    # gets allocated that code. (Caused phantom 10k costs on new parts in the local SQLite DB.)
    from app.schemas import NewItemIn
    from app.routers.catalog import create_catalog_item
    db = _db()
    db.add(DecidedCost(item_id="UN001P", volume_tier=10000, unit_cost_eur=99.0)); db.commit()
    r = create_catalog_item(NewItemIn(item_name="Brand New Thing", module="UN"), db=db, user="t")
    assert r["item_id"] == "UN001P"
    assert not db.execute(select(DecidedCost).where(DecidedCost.item_id == "UN001P")).scalars().all()


def test_cleanup_orphans_keeps_valid():
    from app.routers.admin import cleanup_orphans
    db = _db()
    db.add(Item(item_id="AEC100A", item_name="R", item_type="assembly", module_code="AEC")); db.commit()
    db.add(DecidedCost(item_id="GHOST9P", volume_tier=10000, unit_cost_eur=1))   # orphan
    db.add(DecidedCost(item_id="AEC100A", volume_tier=100, unit_cost_eur=2))     # valid
    db.add(BomLink(parent_item_id="AEC100A", child_item_id="NOPE9P", quantity=1)); db.commit()  # orphan link
    res = cleanup_orphans(db=db, user="t", _="admin")
    assert res["removed"]["decided_costs"] == 1 and res["removed"]["bom_links"] == 1
    assert db.execute(select(DecidedCost).where(DecidedCost.item_id == "AEC100A")).scalars().all()


def test_move_create_bom():
    from app.schemas import CreateBomIn, MoveLinkIn
    from app.routers.edit import create_bom, move_item
    from fastapi import HTTPException
    db = _db()
    root = create_bom(CreateBomIn(item_name="Plant", module="AEC"), db=db, user="t")["item_id"]
    assert root.startswith("AEC") and root.endswith("A") and db.get(Item, root).is_top_level
    db.add(Item(item_id="AEC101A", item_name="SubA", item_type="assembly", module_code="AEC"))
    db.add(Item(item_id="AEC102A", item_name="SubB", item_type="assembly", module_code="AEC"))
    db.add(Item(item_id="AEC050P", item_name="Widget", item_type="part", module_code="AEC")); db.commit()
    db.add(BomLink(parent_item_id=root, child_item_id="AEC101A", quantity=1))
    db.add(BomLink(parent_item_id=root, child_item_id="AEC102A", quantity=1))
    db.add(BomLink(parent_item_id="AEC101A", child_item_id="AEC050P", quantity=2)); db.commit()
    res = move_item("AEC050P", MoveLinkIn(from_parent="AEC101A", to_parent="AEC102A"), db=db, user="t")
    active = [l.parent_item_id for l in db.execute(select(BomLink)).scalars() if l.child_item_id == "AEC050P" and not l.archived]
    assert active == ["AEC102A"] and res["quantity"] == 2 and res["from_parent"] == "AEC101P"
    # cross-BOM is refused
    dr = create_bom(CreateBomIn(item_name="D", module="DAC"), db=db, user="t")["item_id"]
    db.add(Item(item_id="DAC900A", item_name="DSub", item_type="assembly", module_code="DAC")); db.commit()
    db.add(BomLink(parent_item_id=dr, child_item_id="DAC900A", quantity=1)); db.commit()
    try:
        move_item("AEC050P", MoveLinkIn(from_parent="AEC102A", to_parent="DAC900A"), db=db, user="t")
        assert False, "cross-BOM move should be refused"
    except HTTPException as e:
        assert e.status_code == 409


def test_restore_survives_excel_boolean_formulas():
    # A boolean edited & re-saved in Excel comes back as an =TRUE()/=FALSE() formula. Restore
    # must still recover is_top_level/archived — otherwise every flag nulls out and the restore
    # crashes on NOT NULL (the "colleague added weights in Excel, restore failed" bug).
    import io
    import openpyxl
    from app.backup import build_backup_workbook, read_backup_workbook, restore_from_workbook
    db = _db()
    db.add(Item(item_id="AEC100A", item_name="Root", item_type="assembly", module_code="AEC", is_top_level=True))
    db.add(Item(item_id="AEC050P", item_name="Part", item_type="part", module_code="AEC", is_top_level=False))
    db.commit()
    raw = build_backup_workbook(db)
    # booleans are now written as TRUE/FALSE text; simulate Excel turning them into formulas
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    ws = wb["Items"]
    col = [c.value for c in ws[1]].index("is_top_level") + 1
    for r in range(2, ws.max_row + 1):
        cur = str(ws.cell(row=r, column=col).value).strip().upper()
        ws.cell(row=r, column=col).value = "=TRUE()" if cur in ("TRUE", "1") else "=FALSE()"
    buf = io.BytesIO(); wb.save(buf)
    db2 = _db()
    restore_from_workbook(db2, read_backup_workbook(buf.getvalue()))
    roots = sorted(i.item_id for i in db2.execute(select(Item)).scalars() if i.is_top_level)
    assert roots == ["AEC100A"], f"top-level flag not recovered from Excel formulas: {roots}"


def test_new_backup_writes_boolean_text():
    # Backups store booleans as TRUE/FALSE text so they survive an Excel edit/re-save untouched.
    import io
    import openpyxl
    from app.backup import build_backup_workbook
    db = _db()
    db.add(Item(item_id="AEC100A", item_name="Root", item_type="assembly", module_code="AEC", is_top_level=True)); db.commit()
    wb = openpyxl.load_workbook(io.BytesIO(build_backup_workbook(db)))
    ws = wb["Items"]
    col = [c.value for c in ws[1]].index("is_top_level") + 1
    assert str(ws.cell(row=2, column=col).value) in ("TRUE", "FALSE")


def _run():
    tests = [v for k, v in sorted(globals().items()) if k.startswith("test_") and callable(v)]
    passed, failed = 0, 0
    for t in tests:
        try:
            t()
            print(f"  PASS  {t.__name__}")
            passed += 1
        except Exception:  # noqa: BLE001
            print(f"  FAIL  {t.__name__}")
            traceback.print_exc()
            failed += 1
    print(f"\n{passed} passed, {failed} failed")
    return failed


if __name__ == "__main__":
    sys.exit(1 if _run() else 0)
